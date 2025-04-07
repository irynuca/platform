import os
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

INPUT_DIR = r"C:\Users\irina\Project Element\Data source\AQ\AQ_raw\AQ_raw_tables"
OUTPUT_DIR = os.path.join(INPUT_DIR, "unmerged")
os.makedirs(OUTPUT_DIR, exist_ok=True)

def unmerge_sheet(ws_source: Worksheet, ws_target: Worksheet):
    # 1. Unmerge the source sheet (in-place)
    for merged_range in list(ws_source.merged_cells.ranges):
        ws_source.unmerge_cells(str(merged_range))

    # 2. Copy values into the new target sheet
    for row_idx, row in enumerate(ws_source.iter_rows(values_only=True), start=1):
        for col_idx, val in enumerate(row, start=1):
            ws_target.cell(row=row_idx, column=col_idx, value=val)

def process_excel_file(file_path):
    print(f"📂 Unmerging: {os.path.basename(file_path)}")
    wb_source = load_workbook(file_path)
    wb_unmerged = Workbook()
    wb_unmerged.remove(wb_unmerged.active)

    for sheetname in wb_source.sheetnames:
        ws_source = wb_source[sheetname]
        ws_target = wb_unmerged.create_sheet(title=sheetname)
        unmerge_sheet(ws_source, ws_target)

    # Save unmerged version
    unmerged_name = f"{os.path.splitext(os.path.basename(file_path))[0]}_unmerged.xlsx"
    output_path = os.path.join(OUTPUT_DIR, unmerged_name)
    wb_unmerged.save(output_path)
    print(f"✅ Saved unmerged file to: {output_path}")

def batch_process_folder(folder_path):
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            full_path = os.path.join(folder_path, filename)
            process_excel_file(full_path)

if __name__ == "__main__":
    batch_process_folder(INPUT_DIR)
