import os
import re
import sqlite3
from datetime import datetime

import pandas as pd
import unicodedata
from openpyxl import load_workbook
from rapidfuzz import fuzz, process

# --- CONFIG ---
INPUT_DIR = r"C:\Users\irina\Project Element\Data source\AQ\AQ_raw\AQ_clean_tables"
DB_PATH = r"C:\Irina\Mosaiq8\app\data\financials.db"

def remove_diacritics(text):
    replacements = {
        'ă': 'a', 'â': 'a', 'î': 'i', 'ș': 's', 'ş': 's', 'ț': 't', 'ţ': 't',
        'Ă': 'A', 'Â': 'A', 'Î': 'I', 'Ș': 'S', 'Ş': 'S', 'Ț': 'T', 'Ţ': 'T'
    }
    return ''.join(replacements.get(c, c) for c in text)


def load_metrics_from_db(conn):
    df = pd.read_sql_query("SELECT metric_name_ro FROM financial_metrics", conn)
    return df["metric_name_ro"].tolist()


def fuzzy_find_or_insert_metric(conn, raw_name, ticker, threshold=90):
    cleaned_name = remove_diacritics(raw_name.strip())
    cleaned_name_lower = cleaned_name.lower()
    
    existing = load_metrics_from_db(conn)
    match, score, _ = process.extractOne(cleaned_name_lower, existing, scorer=fuzz.token_sort_ratio)

    if score >= threshold:
        print(f"✅ Matched '{raw_name}' to '{match}' (score={score})")
        return match

    print(f"🆕 New metric detected: '{raw_name}' → inserted as '{cleaned_name}' for ticker '{ticker}'")
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO financial_metrics (metric_name_ro, company_ticker)
        VALUES (?, ?)
    """, (cleaned_name, ticker))
    conn.commit()

    return cleaned_name




def get_metric_parent(conn, metric_name_ro):
    cursor = conn.cursor()
    cursor.execute("""
        SELECT metric_parent FROM financial_metrics WHERE metric_name_ro = ?
    """, (metric_name_ro,))
    result = cursor.fetchone()
    return result[0] if result else None

ro_months = {
    "ianuarie": "jan", "ian": "jan", "februarie": "feb", "feb": "feb",
    "martie": "mar", "mar": "mar", "aprilie": "apr", "apr": "apr", "mai": "may",
    "iunie": "jun", "iun": "jun", "iulie": "jul", "iul": "jul", "august": "aug", "aug": "aug",
    "septembrie": "sep", "sept": "sep", "sep": "sep", "octombrie": "oct", "oct": "oct",
    "noiembrie": "nov", "noi": "nov", "decembrie": "dec", "dec": "dec"
}

def normalize_period_string(date_str):
    if isinstance(date_str, datetime):
        return date_str.strftime("%d/%m/%Y")
    if not isinstance(date_str, str):
        return None
    date_str = re.sub(r"[./\s]+", "-", date_str.strip().lower())
    for ro, en in ro_months.items():
        date_str = re.sub(rf"\b{ro}\b", en, date_str)
    formats = ["%d-%b-%Y", "%d-%b-%y", "%d-%m-%Y", "%d-%m-%y", "%d-%B-%Y", "%d-%B-%y"]
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            continue
    return None  # safer than returning invalid string

def extract_date_from_filename(filename):
    # Example: AQ_conso_ro_31.12.2021.xlsx → 31.12.2021
    match = re.search(r"(\d{2})[.\-](\d{2})[.\-](\d{4})", filename)
    if match:
        try:
            return datetime.strptime(match.group(0), "%d.%m.%Y")
        except ValueError:
            pass
    return datetime.min  # fallback for bad formats

def parse_value(cell):
    """Convert cell to float, treating placeholders like '–', '.' or '' as 0."""
    if pd.isna(cell):
        return None
    if isinstance(cell, str):
        cleaned = cell.strip().replace("–", "").replace(",", ".").replace("−", "-")
        if cleaned in ["", ".", "-"]:
            return 0.0
        try:
            return float(cleaned)
        except ValueError:
            return None
    try:
        return float(cell)
    except (ValueError, TypeError):
        return None


def extract_cf_with_fuzzy_metrics(filepath, conn):
    filename = os.path.basename(filepath)
    company_ticker = filename.split("_")[0]
    statement_type = "consolidated"
    statement_name = "CashFlow"

    print(f"\n📄 Processing file: {filename}")

    wb = load_workbook(filepath, data_only=True)
    for sheetname in wb.sheetnames:
        if not re.match(r"CF_\d{8}", sheetname):
            continue

        print(f"🔍 Reading sheet: {sheetname}")
        df = pd.read_excel(filepath, sheet_name=sheetname)
        df.columns = [normalize_period_string(col) for col in df.columns]

        for idx, row in df.iterrows():
            raw_metric = str(row.iloc[0]).strip()
            cleaned_metric = remove_diacritics(raw_metric)
            
            # Skip invalid rows
            val1, val2 = row.iloc[1], row.iloc[2]
            if (pd.isna(val1) or val1 == '') and (pd.isna(val2) or val2 == ''):
                continue
            val1 = parse_value(row.iloc[1])
            val2 = parse_value(row.iloc[2])

            # Skip if still both are None (e.g., completely invalid row)
            if val1 is None and val2 is None:
                print(f"⚠️ Skipping non-numeric row at line {idx+2}: '{raw_metric}'")
                continue

            matched_metric = fuzzy_find_or_insert_metric(conn, cleaned_metric, company_ticker)
            print(f"🧼 Cleaned metric: '{raw_metric}' → '{cleaned_metric}' → using: '{matched_metric}'")

            col_data = [
                (df.columns[1], val1),
                (df.columns[2], val2)
            ]

            for period_label, value in col_data:
                if value is None or not period_label:
                    continue

                period_label_norm = normalize_period_string(period_label)
                if not period_label_norm:
                    continue

                try:
                    period_end = datetime.strptime(period_label_norm, "%d/%m/%Y")
                except ValueError:
                    print(f"⚠️ Invalid period: '{period_label}'")
                    continue

                period_start = datetime(period_end.year, 1, 1)
                this_period_type = "annual" if period_label_norm.startswith("31/12") else "quarter"

                print(f"➡️ Line {idx+2}: {matched_metric} = {value} on {period_label_norm}")

                record = {
                    "company_ticker": company_ticker,
                    "statement_name": statement_name,
                    "statement_type": statement_type,
                    "period_start": period_start.strftime("%Y-%m-%d"),
                    "period_end": period_end.strftime("%Y-%m-%d"),
                    "period_type": this_period_type,
                    "aggr_type": "cml",
                    "currency": "RON",
                    "metric_name_ro": matched_metric,
                    "value": value,
                    "metric_parent": None,
                    "line_order": idx + 1,
                    "last_updated": datetime.now().strftime("%Y-%m-%d")
                }

                placeholders = ", ".join(record.keys())
                values = tuple(record.values())
                sql = f"""
                    INSERT OR REPLACE INTO cashflow_staging ({placeholders})
                    VALUES ({','.join(['?'] * len(record))})
                """
                conn.execute(sql, values)

        conn.commit()
        print(f"✅ Finished sheet: {sheetname}")


def process_all_excels(directory, conn):
    excel_files = [
        os.path.join(directory, f)
        for f in os.listdir(directory)
        if f.endswith(".xlsx")
    ]

    sorted_files = sorted(excel_files, key=extract_date_from_filename)

    for file_path in sorted_files:
        extract_cf_with_fuzzy_metrics(file_path, conn)

def export_cashflow_series_to_excel(conn, output_path="cashflow_series.xlsx"):
    print("📊 Exporting cash flow series to Excel...")

    query = """
    SELECT 
        company_ticker,
        period_end,
        metric_name_ro,
        value,
        line_order
    FROM cashflow_staging
    ORDER BY line_order ASC, metric_name_ro, period_end
    """

    df = pd.read_sql_query(query, conn)

    if df.empty:
        print("⚠️ No data found in cashflow_staging.")
        return

    # Maintain line order
    df["metric_name_ro"] = pd.Categorical(df["metric_name_ro"], 
                                          categories=df.sort_values("line_order")["metric_name_ro"].unique(),
                                          ordered=True)

    pivot_df = df.pivot_table(
        index="metric_name_ro",
        columns="period_end",
        values="value",
        aggfunc="first"
    )

    # Format period columns
    pivot_df.columns = pd.to_datetime(pivot_df.columns).strftime('%d-%m-%Y')
    pivot_df = pivot_df[sorted(pivot_df.columns, key=lambda x: pd.to_datetime(x, dayfirst=True))]

    # Save to Excel
    excel_path = os.path.join(os.getcwd(), output_path)
    pivot_df.to_excel(excel_path)
    print(f"✅ Cash flow series exported to {excel_path}")


def init_db(db_path):
    conn = sqlite3.connect(db_path)
    return conn

# --- Run ---
if __name__ == "__main__":
    print("🚀 Starting cash flow data extraction...")
    conn = init_db(DB_PATH)
    try:
        process_all_excels(INPUT_DIR, conn)
        export_cashflow_series_to_excel(conn)
    finally:
        conn.close()
    print("✅ Done: All CF data processed and exported.")
