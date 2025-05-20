import os
import re
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from rapidfuzz import fuzz, process

# --- CONFIG ---
DB_PATH = r"C:\Irina\Mosaiq8\app\data\financials.db"
BASE_DIR_TEMPLATE = r"C:\Users\irina\Project Element\Data source\{ticker}\{ticker}_raw\{ticker}_clean_tables"
GLOBAL_MATCHING_THRESHOLD = 95
METRIC_MAPPING_FILE = "metric_mapping.json"

# --- UTILITIES ---

def get_user_inputs():
    ticker = input("Enter the company ticker (e.g., AQ): ").strip().upper()
    statement = input("Enter the statement type (CF, BS, PL): ").strip().upper()
    if statement not in ["CF", "BS", "PL"]:
        print(f"Invalid statement '{statement}'. Choose CF, BS, or PL.")
        exit(1)
    input_dir = BASE_DIR_TEMPLATE.format(ticker=ticker)
    if not os.path.isdir(input_dir):
        print(f"Directory not found: {input_dir}")
        exit(1)
    print(f"📂 Using input directory: {input_dir}")
    return ticker, statement, input_dir


def remove_diacritics(text):
    mapping = {'ă':'a','â':'a','î':'i','ș':'s','ş':'s','ț':'t','ţ':'t',
               'Ă':'A','Â':'A','Î':'I','Ș':'S','Ş':'S','Ț':'T','Ţ':'T'}
    return ''.join(mapping.get(c, c) for c in text)


def clean_metric_name(raw):
    no_diac = remove_diacritics(raw.strip().lower())
    return re.sub(r'[^a-z0-9 ]+', '', no_diac)


def is_valid_line(metric, vals):
    if not metric.strip():
        return False
    for v in vals:
        if pd.notna(v) and str(v).strip() not in ['', '-', '0', '0.0', 'nan']:
            return True
    return False

# --- MAPPING ---
metric_mapping = {}


def load_metric_mappings():
    if os.path.exists(METRIC_MAPPING_FILE):
        print("🔄 Loading existing metric mappings...")
        return json.load(open(METRIC_MAPPING_FILE, 'r', encoding='utf-8'))
    print("⚠️ Mapping file not found, creating new.")
    json.dump({}, open(METRIC_MAPPING_FILE,'w',encoding='utf-8'), ensure_ascii=False, indent=4)
    return {}


def save_metric_mappings():
    json.dump(metric_mapping, open(METRIC_MAPPING_FILE,'w',encoding='utf-8'), ensure_ascii=False, indent=4)


def fuzzy_find_or_insert_metric(raw_name, values, interactive=True, threshold=GLOBAL_MATCHING_THRESHOLD):
    if not is_valid_line(raw_name, values):
        print(f"⚠️ Skipping invalid line: '{raw_name}' with values {values}")
        return None
    cleaned = clean_metric_name(raw_name)
    # if known
    if cleaned in metric_mapping:
        return metric_mapping[cleaned]
    # first file: seed silently
    if not interactive:
        metric_mapping[cleaned] = cleaned
        return cleaned
    # try fuzzy match
    existing = list(metric_mapping.values())
    if existing:
        match, score, _ = process.extractOne(cleaned, existing, scorer=fuzz.token_sort_ratio)
        if score >= threshold:
            ans = input(f"\n🆕 New metric '{raw_name}'. Replace with '{match}'? (y/n): ").strip().lower()
            if ans == 'y':
                metric_mapping[cleaned] = match
                save_metric_mappings()
                return match
    # no auto-match
    ans = input(f"\n🆕 Metric '{raw_name}' new. Add as new (y) or match existing (n)? ").strip().lower()
    if ans == 'y':
        metric_mapping[cleaned] = cleaned
        save_metric_mappings()
        return cleaned
    # propose all matches
    possible = process.extract(cleaned, existing, scorer=fuzz.token_sort_ratio, limit=len(existing))
    print("\n🔄 Possible Matches:")
    for i, (m, s, _) in enumerate(possible, 1):
        print(f"{i}. {m} (score={s})")
    while True:
        choice = input("Choose number or enter new name: ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(possible):
            sel = possible[int(choice)-1][0]
            metric_mapping[cleaned] = sel
            save_metric_mappings()
            return sel
        if choice:
            metric_mapping[cleaned] = choice
            save_metric_mappings()
            return choice
        print("Invalid choice.")

# --- DATE ---
ro_months = {
    'ianuarie':'jan','ian':'jan','februarie':'feb','feb':'feb',
    'martie':'mar','mar':'mar','aprilie':'apr','apr':'apr','mai':'may',
    'iunie':'jun','iun':'jun','iulie':'jul','iul':'jul','august':'aug','aug':'aug',
    'septembrie':'sep','sept':'sep','sep':'sep','octombrie':'oct','oct':'oct',
    'noiembrie':'nov','noi':'nov','decembrie':'dec','dec':'dec'
}


def normalize_period_string(ds):
    if isinstance(ds, datetime):
        return ds.strftime('%d/%m/%Y')
    if not isinstance(ds, str):
        return None
    s = re.sub(r'[./\s]+','-', ds.strip().lower())
    for ro,en in ro_months.items(): s = re.sub(rf"\b{ro}\b", en, s)
    formats = ['%d-%b-%Y','%d-%b-%y','%d-%m-%Y','%d-%m-%y','%Y-%m-%d','%b-%d-%Y','%d-%B-%Y','%d/%m/%Y','%d.%m.%Y']
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).strftime('%d/%m/%Y')
        except:
            continue
    print(f"⚠️ Could not parse date: '{ds}'")
    return None


def extract_date_from_filename(fn):
    m = re.search(r'(\d{2})[.\-](\d{2})[.\-](\d{4})', fn)
    if m:
        try:
            return datetime.strptime(m.group(0), '%d.%m.%Y')
        except:
            pass
    return datetime.min


def parse_value(c):
    if pd.isna(c): return None
    if isinstance(c,str):
        cl = c.strip().replace(',', '').replace('–','').replace('−','-')
        if cl in ['','.','-']: return 0.0
        try: return float(cl)
        except: return None
    try: return float(c)
    except: return None

# --- EXTRACT ---
def extract_data_to_list(fp, stmt, interactive=True):
    wb = load_workbook(fp, data_only=True)
    records = []
    for sh in wb.sheetnames:
        if not re.match(rf"{stmt}_\d{{8}}", sh): continue
        df = pd.read_excel(fp, sheet_name=sh)
        df.columns = [df.columns[0]] + [normalize_period_string(c) for c in df.columns[1:]]
        for idx,row in df.iterrows():
            raw = str(row.iloc[0])
            vals = [row.iloc[1], row.iloc[2]]
            norm = fuzzy_find_or_insert_metric(raw, vals, interactive)
            if not norm: continue
            for col,val in zip(df.columns[1:], vals):
                if pd.isna(val): continue
                dnorm = normalize_period_string(col)
                if not dnorm: continue
                records.append({'metric_name_ro':norm,'period_end':dnorm,'value':parse_value(val)})
    return records

# --- MAIN ---
if __name__ == '__main__':
    ticker, stmt, inp = get_user_inputs()
    metric_mapping.clear()
    metric_mapping.update(load_metric_mappings())
    all_records = []
    files = sorted([os.path.join(inp,f) for f in os.listdir(inp) if f.endswith('.xlsx')], key=extract_date_from_filename)
    for i,fp in enumerate(files):
        interactive = (i > 0)
        print(f"📄 Processing {'first' if i==0 else 'subsequent'} file: {os.path.basename(fp)}")
        recs = extract_data_to_list(fp, stmt, interactive)
        all_records.extend(recs)
    df = pd.DataFrame(all_records)
    if not df.empty:
        order = df['metric_name_ro'].drop_duplicates().tolist()
        pivot = df.pivot_table(index='metric_name_ro', columns='period_end', values='value', aggfunc='first')
        pivot = pivot.reindex(order)
        cols = sorted(pivot.columns, key=lambda x: datetime.strptime(x, '%d/%m/%Y'))
        pivot = pivot[cols]
        out = f"{ticker}_{stmt}_extracted.xlsx"
        pivot.to_excel(out)
        print(f"✅ Exported to {out}")
    save_metric_mappings()
