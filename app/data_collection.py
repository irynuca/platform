import os
import re
import sqlite3
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from rapidfuzz import fuzz

# --- CONFIG ---
INPUT_DIR = r"C:\Users\irina\Project Element\Data source\AQ\AQ_raw\AQ_clean_tables"
DB_PATH = r"C:\Irina\Mosaiq8\app\data\financials.db"

# --- Setup: Month translation + date normalization ---
ro_months = {
    "ianuarie": "jan", "ian": "jan", "februarie": "feb", "feb": "feb",
    "martie": "mar", "mar": "mar", "aprilie": "apr", "apr": "apr", "mai": "may",
    "iunie": "jun", "iun": "jun", "iulie": "jul", "iul": "jul", "august": "aug", "aug": "aug",
    "septembrie": "sep", "sept": "sep", "sep": "sep", "octombrie": "oct", "oct": "oct",
    "noiembrie": "nov", "noi": "nov", "decembrie": "dec", "dec": "dec"
}

def normalize_period_string(date_str):
    if isinstance(date_str, datetime):
        return date_str.strftime("%d/%m/%Y")
    if not isinstance(date_str, str):
        return None
    date_str = re.sub(r"[./\s]+", "-", date_str.strip().lower())
    for ro, en in ro_months.items():
        date_str = re.sub(rf"\b{ro}\b", en, date_str)
    formats = ["%d-%b-%Y", "%d-%b-%y", "%d-%m-%Y", "%d-%m-%y", "%d-%B-%Y", "%d-%B-%y"]
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            continue
    return None  # safer than returning invalid string


def extract_parent_metric(sheetname):
    match = re.match(r"notes_(.*?)_\d{8}$", sheetname.lower())
    return match.group(1) if match else "unknown"

# --- DB Init ---
def init_db(db_path):
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS notes (
    note_id INTEGER PRIMARY KEY AUTOINCREMENT,
    company_ticker TEXT NOT NULL,
    period_end DATE NOT NULL,
    period_type TEXT NOT NULL,
    aggr_type TEXT NOT NULL,
    note_type TEXT,
    parent_metric TEXT,
    note_element TEXT NOT NULL,
    value REAL NOT NULL,
    currency TEXT DEFAULT 'RON',
    line_order INTEGER DEFAULT 0,
    source TEXT,
    content TEXT,
    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
    UNIQUE (company_ticker, period_end, note_element)
);
    """)
    conn.commit()
    return conn

def insert_note_record(conn, data: dict):
    placeholders = ", ".join(data.keys())
    columns = list(data.keys())
    values = tuple(data.values())

    sql = f"""
    INSERT OR REPLACE INTO notes ({placeholders})
    VALUES ({','.join(['?'] * len(columns))})
    """
    conn.execute(sql, values)


# --- Fuzzy Match Setup ---
target_keywords = ["Cheltuieli cu dobanzile"]

def is_fuzzy_match(text, keywords, threshold=85):
    return any(fuzz.partial_ratio(text.lower(), k.lower()) >= threshold for k in keywords)

# --- Main extraction ---
def extract_interest_expense_from_file(filepath, conn):
    filename = os.path.basename(filepath)
    wb = load_workbook(filepath, data_only=True)

    for sheetname in wb.sheetnames:
        if re.match(r"notes_.*_\d{8}", sheetname.lower()):
            df = pd.read_excel(filepath, sheet_name=sheetname, header=0)
            df.columns = [normalize_period_string(c) for c in df.columns]

            for idx, row in df.iterrows():
                metric = str(row.iloc[0]).strip()
                if is_fuzzy_match(metric, target_keywords):
                    try:
                        col_data = [
                            (normalize_period_string(df.columns[1]), float(row.iloc[1])),
                            (normalize_period_string(df.columns[2]), float(row.iloc[2]))
                        ]
                    except Exception as e:
                        print(f"⚠️ Skipping row due to parse error: {e}")
                        continue

                    parent_metric = extract_parent_metric(sheetname)
                    print(f"\n📄 {filename} | Sheet: {sheetname}")
                    print(f"🧾 Metric: {metric}")

                    for i, (period_label, value) in enumerate(col_data, start=1):
                        if period_label is None:
                            print(f"⚠️ Skipping entry due to invalid period label in: {filename}")
                            continue

                        this_period_type = "annual" if period_label.startswith("31/12") else "quarter"


                        record = {
                            "company_ticker": filename.split("_")[0],
                            "period_end": period_label,
                            "period_type": this_period_type,
                            "aggr_type": "cml",
                            "note_element": target_keywords[0],
                            "value": value,
                            "line_order": idx + 1,
                            "currency": "RON",
                            "parent_metric": parent_metric,
                            "note_type": parent_metric,
                            "source": filename,
                            "content": None
                        }
                        insert_note_record(conn, record)
                    conn.commit()
                    return  # Stop after first match

# --- Batch runner ---
def extract_date_from_filename(filename):
    # Example: AQ_conso_ro_31.12.2021.xlsx → 31.12.2021
    match = re.search(r"(\d{2})[.\-](\d{2})[.\-](\d{4})", filename)
    if match:
        try:
            return datetime.strptime(match.group(0), "%d.%m.%Y")
        except ValueError:
            pass
    return datetime.min  # fallback for bad formats

def process_all_excels(directory, conn):
    excel_files = [
        os.path.join(directory, f)
        for f in os.listdir(directory)
        if f.endswith(".xlsx")
    ]

    # Sort files by the embedded date in their filename
    sorted_files = sorted(excel_files, key=extract_date_from_filename)
    print(sorted_files)
    for file_path in sorted_files:
        extract_interest_expense_from_file(file_path, conn)


# --- Run ---
if __name__ == "__main__":
    conn = init_db(DB_PATH)
    process_all_excels(INPUT_DIR, conn)
    conn.close()
    print("✅ Done: All matched notes inserted into the database.")
