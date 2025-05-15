import os
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import re

def clean_excel_files(folder_path):
    for file_name in os.listdir(folder_path):
        if file_name.endswith(('.xlsx', '.xlsm', '.xls')):
            file_path = os.path.join(folder_path, file_name)
            print(f"Processing: {file_path}")
            workbook = openpyxl.load_workbook(file_path)
            # Extract date from file name (e.g., 12.05.2025 -> 12052025)
            date_match = re.search(r'\d{2}\.\d{2}\.\d{4}', file_name)
            if date_match:
                date_str = date_match.group().replace('.', '')
            else:
                print(f"Warning: No date found in file name {file_name}, skipping.")
                continue

            # Check for keywords to identify financial sheets
            bs_keywords = ["Capitaluri", "Imobilizari", "Active", "Datorii"]
            pl_keywords = ["Venituri", "Cheltuieli", "Impozit", "Profit"]
            bs_keywords = ["Capitaluri", "Imobilizari", "Active", "Datorii"]
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                df = pd.DataFrame(sheet.values).fillna('')

                # Check for distinct rows containing each required keyword
                # Check for BS keywords
                bs_keyword_hits = set()
                for row in df.values:
                    row_text = ' '.join(map(str, row))
                    for keyword in bs_keywords:
                        if keyword in row_text:
                            bs_keyword_hits.add(keyword)
                            break  # Stop checking this row once a keyword is matched
                    # If all BS keywords are found in distinct rows, rename the sheet
                    if len(bs_keyword_hits) == len(bs_keywords):
                        new_sheet_name = f"BS_{date_str}"
                        print(f"Renaming sheet {sheet_name} to {new_sheet_name}")
                        sheet.title = new_sheet_name
                        break
                
                # Check for PL keywords
                pl_keyword_hits = set()
                for row in df.values:
                    row_text = ' '.join(map(str, row))
                    for keyword in pl_keywords:
                        if keyword in row_text:
                            pl_keyword_hits.add(keyword)
                            break  # Stop checking this row once a keyword is matched
                    # If all PL keywords are found in distinct rows, rename the sheet
                    if len(pl_keyword_hits) == len(pl_keywords):
                        new_sheet_name = f"PL_{date_str}"
                        print(f"Renaming sheet {sheet_name} to {new_sheet_name}")
                        sheet.title = new_sheet_name
                        break
                keyword_hits = set()
                for row in df.values:
                    row_text = ' '.join(map(str, row))
                    for keyword in bs_keywords:
                        if keyword in row_text:
                            keyword_hits.add(keyword)
                            break  # Stop checking this row once a keyword is matched
                    # If all 4 keywords are found in distinct rows, rename the sheet
                    if len(keyword_hits) == len(bs_keywords):
                        new_sheet_name = f"BS_{date_str}"
                        print(f"Renaming sheet {sheet_name} to {new_sheet_name}")
                        sheet.title = new_sheet_name
                        break

                # 1. Unmerge cells
                if sheet.merged_cells.ranges:
                    for merged_range in list(sheet.merged_cells.ranges):
                        sheet.unmerge_cells(str(merged_range))

                # 2. Unwrap text and 3. Delete empty rows and columns
                df.dropna(how='all', axis=0, inplace=True)  # Remove empty rows
                df.dropna(how='all', axis=1, inplace=True)  # Remove empty columns

                # Clear the original sheet and rewrite the cleaned data
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.value = None

                for r_idx, row in df.iterrows():
                    for c_idx, value in enumerate(row):
                        cell_ref = f"{get_column_letter(c_idx + 1)}{r_idx + 1}"
                        sheet[cell_ref].value = value

            # Save the cleaned workbook
            workbook.save(file_path)
            print(f"Cleaned and saved: {file_path}\n")

# Run the script for a given folder
folder_path = r"C:\Users\irina\Project Element\Data source\WINE\WINE_raw\WINE_clean_tables"  # Replace with your target folder path
clean_excel_files(folder_path)
